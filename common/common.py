from base_page.base_page import BasePage
from selenium.common.exceptions import NoSuchElementException
from appium.webdriver.common.touch_action import TouchAction
from appium.webdriver.common.multi_action import MultiAction
from init.init import start_app
from selenium.webdriver.common.by import By
from log.log import Logger
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.header import Header


import openpyxl
class Common(BasePage):
    #取消按钮
    cance_bt = (By.ID,"android:id/button2")
    #跳过按钮
    skip_bt = (By.ID,"com.tal.kaoyan:id/tv_skip")

    def cance(self):
        Logger().log().info("检查是否有取消按钮")
        try:
            cance = self.loc(self.cance_bt)
        except NoSuchElementException:
            Logger().log().info("没有取消按钮")
        else:
            Logger().log().info("点击取消按钮")
            cance.click()

    def skip(self):
        Logger().log().info("检查是否有跳过按钮")
        try:
            skip = self.loc(self.skip_bt)
        except NoSuchElementException:
            Logger().log().info("没有跳过按钮")
        else:
            skip.click()
            Logger().log().info("点击跳过按钮")

    def size(self):
        x = self.driver.get_window_size()['width']
        y = self.driver.get_window_size()['height']
        return x, y

    def slide(self, x1, y1, x2, y2):
        size = Common(self.driver).size()
        x1 = int(size[0]*x1)
        y1 = int(size[1]*y1)
        x2 = int(size[0]*x2)
        y2 = int(size[1]*y2)
        self.driver.swipe(x1, y1, x2, y2, 1000)

    # def left_slip(self):
    #     size = Common(self.driver).size()
    #     x1 = int(size[0]*0.9)
    #     x2 = int(size[0]*0.1)
    #     y = int(size[1]*0.5)
    #     self.driver.swipe(x1, y, x2, y, 1000)
    #
    # def right_slip(self):
    #     size = Common(self.driver).size()
    #     x1 = int(size[0]*0.1)
    #     x2 = int(size[0]*0.9)
    #     y = int(size[1]*0.5)
    #     self.driver.swipe(x1, y, x2, y, 1000)
    #
    # def up_slip(self):
    #     size = Common(self.driver).size()
    #     x = int(size[0]*0.5)
    #     y1 = int(size[1]*0.9)
    #     y2 = int(size[1]*0.1)
    #     self.driver.swipe(x, y1, x, y2, 1000)
    #
    # def down_slip(self):
    #     size = Common(self.driver).size()
    #     x = int(size[0]*0.5)
    #     y1 = int(size[1]*0.1)
    #     y2 = int(size[1]*0.9)
    #     self.driver.swipe(x, y1, x, y2, 1000)

    #连续滑动解锁
    def slide_unlock(self):
        TouchAction(self.driver)\
            .press(x=246, y=376).wait(2000) \
            .move_to(x=451, y=385).wait(1000) \
            .move_to(x=644, y=584).wait(1000) \
            .move_to(x=661, y=774).wait(1000) \
            .release().perform()


    def zoom(self, x1, y1, x2, y2, a1, b1, a2, b2):
        size = Common(self.driver).size()
        action1 = TouchAction(self.driver)
        action2 = TouchAction(self.driver)
        zoom_action = MultiAction(self.driver)
        action1.press(x=size[0] * x1, y=size[1] * y1).wait(1000).move_to(x=size[0] * x2, y=size[1] * y2).wait(1000).release()
        action2.press(x=size[0] * a1, y=size[1] * b1).wait(1000).move_to(x=size[0] * a2, y=size[1] * b2).wait(1000).release()

        zoom_action.add(action1, action2)
        zoom_action.perform()

#发送邮件
def send_email(email_path):
    message = MIMEMultipart()
    #邮件内容
    text = """
    请输入你想说的邮件内容
    """
    message.attach(MIMEText(_text=text, _subtype='plain', _charset="utf-8"))
    #需要发送的附件的路径
    with open(email_path, 'rb') as f:
        content = f.read()
    att1 = MIMEText(content, "base64", "utf-8")
    att1["Content-Type"] = 'application/octet-stream'
    att1['Content-Disposition'] = 'attachment; filename = "report.html"'
    message.attach(att1)

    #邮件主题
    message["Subject"] = Header("主题", "utf-8").encode()
    message["From"] = Header("tianmeng", "utf-8")
    message["To"] = Header('tianmeng_wxk', "utf-8")

    try:
        smtp = smtplib.SMTP()
        #smtp = smtplib.SMTP_SSL('smtp.163.com', 465)
        smtp.connect(host="smtp.qq.com", port=587)
        smtp.login(user="3394788013@qq.com", password="lizceyidpekpdbhd")
        sender = "3394788013@qq.com"
        receiver = ['tianmeng_wxk@163.com']
        smtp.sendmail(sender, receiver, message.as_string())
        Logger().log().info("发送邮件成功")
        return email_path
    except smtplib.SMTPException as e:
        Logger().log().info("发送邮件失败，失败信息：{}".format(e))


# openpyxl加载excel
def load_excel(excel_path):
    global excel
    excel = openpyxl.load_workbook(excel_path)
    return excel

# openpyxl加载sheet
def load_sheet(sheet_name):
    sheet = excel[sheet_name]
    return sheet

# openpyxl读取数据
def read_excel(excel_path, sheet_name):
    excel = load_excel(excel_path)
    sheet = load_sheet(sheet_name)
    l = []
    for i in range(2, sheet.max_row + 1):
        data = []
        arg1 = sheet.cell(i, 1).value
        arg2 = sheet.cell(i, 2).value
        #arg3 = sheet.cell(i, 3).value
        data.append(arg1)
        data.append(arg2)
        #data.append(arg3)
        l.append(data)
    return l


if __name__ == '__main__':
    driver = start_app()
    com = Common(driver)
    com.cance()
    com.skip()